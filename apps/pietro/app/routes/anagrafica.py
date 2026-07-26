import io
import json
import re
import sys
from pathlib import Path
from urllib import error as urllib_error
from urllib import request as urllib_request
from datetime import date, datetime, timezone
from flask import Blueprint, current_app, g, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import joinedload, selectinload
from ..models import db, Iscritto, Contabilita, PagamentoSettimanale
from ..config_manager import load_config
from ..request_context import get_current_event_id, get_iscritto_or_404

# Import delle funzioni del report standalone (genera_report.py nella root del repo)
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
import genera_report as _rpt  # noqa: E402
import assegna_squadre as _aq  # noqa: E402
import random as _random
import secrets as _secrets
from collections import defaultdict as _defaultdict

bp = Blueprint("anagrafica", __name__)

_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
_PRESENZA_FIELDS = {f"PresenzaSettimana{i}" for i in range(1, 13)}
_MOBILE_UA_HINTS = ("iphone", "android", "mobile", "ipad", "ipod")


def _validate_iscritto(data):
    """Return a list of validation error strings (empty means valid)."""
    errors = []
    if not isinstance(data, dict):
        return ["Payload JSON non valido."]

    # Required fields
    nome = data.get("NomeRagazzo", "").strip() if data.get("NomeRagazzo") else ""
    cognome = data.get("CognomeRagazzo", "").strip() if data.get("CognomeRagazzo") else ""
    if not nome:
        errors.append("NomeRagazzo e' obbligatorio.")
    if not cognome:
        errors.append("CognomeRagazzo e' obbligatorio.")

    # Email validation
    for field in ("MailMamma", "MailPapa", "MailRicevuta"):
        val = data.get(field, "")
        if val and val.strip():
            if not _EMAIL_RE.match(val.strip()):
                errors.append(f"{field} non e' un indirizzo email valido.")

    for field in ("DataNascitaRagazzo", "DataValidazione"):
        raw = data.get(field)
        if raw:
            try:
                _parse_iso_date(raw)
            except ValueError:
                errors.append(f"{field} non e' una data valida.")

    return errors


def _current_actor():
    auth_user = getattr(g, "auth_user", None) or {}
    return auth_user.get("name") or "App"


def _parse_iso_date(raw_value):
    if raw_value in (None, ""):
        return None
    if isinstance(raw_value, date):
        return raw_value
    return date.fromisoformat(str(raw_value))


def _coerce_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "si", "sì", "yes", "on"}
    return bool(value)


def _request_prefers_mobile():
    if request.args.get("desktop", "").strip().lower() in {"1", "true", "yes"}:
        return False

    user_agent = (request.user_agent.string or "").lower()
    return any(hint in user_agent for hint in _MOBILE_UA_HINTS)


def _presence_by_week(iscritto, numero_settimane):
    presenze = {}
    contabilita = iscritto.contabilita

    if not contabilita:
        for settimana in range(1, numero_settimane + 1):
            presenze[settimana] = False
        return presenze

    if contabilita.Gratuita:
        for settimana in range(1, numero_settimane + 1):
            presenze[settimana] = True
        return presenze

    pagamenti = {
        item.NumeroSettimana: bool(item.Pagato)
        for item in contabilita.settimane
    }
    for settimana in range(1, numero_settimane + 1):
        presenze[settimana] = pagamenti.get(settimana, False)
    return presenze


def _invoke_resend_confirmation_flow(nome, cognome, email):
    payload = json.dumps({
        "nome": nome,
        "cognome": cognome,
        "email": email,
    }).encode("utf-8")
    flow_url = current_app.config["RESEND_CONFIRMATION_FLOW_URL"]
    timeout_seconds = current_app.config["RESEND_CONFIRMATION_FLOW_TIMEOUT_SECONDS"]
    request_obj = urllib_request.Request(
        flow_url,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib_request.urlopen(request_obj, timeout=timeout_seconds) as response:
            response.read()
            return getattr(response, "status", response.getcode())
    except urllib_error.HTTPError as exc:
        response_body = exc.read().decode("utf-8", errors="replace")
        current_app.logger.warning(
            "Errore reinvio email conferma per %s %s <%s>: HTTP %s - %s",
            nome,
            cognome,
            email,
            exc.code,
            response_body,
        )
        raise RuntimeError(f"Il servizio email ha risposto con errore {exc.code}.") from exc
    except urllib_error.URLError as exc:
        current_app.logger.warning(
            "Servizio reinvio email non raggiungibile per %s %s <%s>: %s",
            nome,
            cognome,
            email,
            exc.reason,
        )
        raise RuntimeError("Il servizio email non e' raggiungibile al momento.") from exc


@bp.route("/")
def home():
    if _request_prefers_mobile():
        return redirect(url_for("api_mobile.mobile_page"))

    config = load_config()
    evento = get_current_event_id()
    iscritti = (Iscritto.query
                .filter_by(ID_Evento=evento)
                .order_by(Iscritto.CognomeRagazzo, Iscritto.NomeRagazzo)
                .all())
    return render_template("home.html", iscritti=iscritti, config=config)


@bp.route("/api/iscritti")
def api_lista_iscritti():
    evento = get_current_event_id()
    q = request.args.get("q", "").strip()

    query = Iscritto.query.filter_by(ID_Evento=evento)
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                Iscritto.NomeRagazzo.ilike(like),
                Iscritto.CognomeRagazzo.ilike(like),
            )
        )
    iscritti = query.order_by(Iscritto.CognomeRagazzo, Iscritto.NomeRagazzo).all()

    result = []
    for i in iscritti:
        d = {
            "ID": i.ID,
            "NomeRagazzo": i.NomeRagazzo,
            "CognomeRagazzo": i.CognomeRagazzo,
            "DataNascitaRagazzo": i.DataNascitaRagazzo.isoformat() if i.DataNascitaRagazzo else None,
        }
        result.append(d)
    return jsonify(result)


@bp.route("/api/iscritti/<int:id>")
def api_dettaglio_iscritto(id):
    iscritto = get_iscritto_or_404(id)
    config = load_config()
    numero_settimane = config["settimane"]["numero_settimane"]
    data = {}
    for c in Iscritto.__table__.columns:
        val = getattr(iscritto, c.name)
        if isinstance(val, (date, datetime)):
            val = val.isoformat()
        data[c.name] = val

    for settimana, presente in _presence_by_week(iscritto, numero_settimane).items():
        data[f"PresenzaSettimana{settimana}"] = presente
    return jsonify(data)


@bp.route("/api/iscritti", methods=["POST"])
def api_crea_iscritto():
    data = request.get_json(silent=True)
    config = load_config()
    evento = get_current_event_id()

    # Validation
    errors = _validate_iscritto(data)
    if errors:
        return jsonify({"errors": errors}), 400

    try:
        iscritto = Iscritto(ID_Evento=evento)
        _update_iscritto_fields(iscritto, data)
        iscritto.ModificatoDa = _current_actor()
        db.session.add(iscritto)
        db.session.flush()

        # Crea contabilita e settimane vuote
        cont = Contabilita(
            ID_Iscritto=iscritto.ID, ID_Evento=evento,
            ImportoIscrizione=config["importi_default"]["iscrizione"]
        )
        cont.ModificatoDa = _current_actor()
        db.session.add(cont)
        db.session.flush()

        num_sett = config["settimane"]["numero_settimane"]
        for s in range(1, num_sett + 1):
            ps = PagamentoSettimanale(ID_Contabilita=cont.ID, NumeroSettimana=s)
            ps.ModificatoDa = _current_actor()
            db.session.add(ps)

        db.session.commit()
        return jsonify({"id": iscritto.ID}), 201
    except Exception:
        current_app.logger.exception("Errore durante la creazione dell'iscritto")
        db.session.rollback()
        return jsonify({"error": "Errore creazione iscritto."}), 500


@bp.route("/api/iscritti/<int:id>", methods=["PUT"])
def api_modifica_iscritto(id):
    iscritto = get_iscritto_or_404(id)
    data = request.get_json(silent=True)

    # Validation
    errors = _validate_iscritto(data)
    if errors:
        return jsonify({"errors": errors}), 400

    _update_iscritto_fields(iscritto, data)
    iscritto.DataModifica = datetime.now(timezone.utc)
    iscritto.ModificatoDa = _current_actor()
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/iscritti/<int:id>", methods=["DELETE"])
def api_elimina_iscritto(id):
    iscritto = get_iscritto_or_404(id)
    try:
        db.session.execute(
            db.text(
                """
                UPDATE [dbo].[import_forms_log]
                SET [ID_Iscritto] = NULL
                WHERE [ID_Iscritto] = :id
                """
            ),
            {"id": id},
        )
        db.session.delete(iscritto)
        db.session.commit()
        return jsonify({"ok": True})
    except Exception:
        current_app.logger.exception("Errore durante l'eliminazione dell'iscritto")
        db.session.rollback()
        return jsonify({"error": "Errore eliminazione iscritto."}), 500


@bp.route("/api/iscritti/<int:id>/reinvia-email-conferma", methods=["POST"])
def api_reinvia_email_conferma(id):
    iscritto = get_iscritto_or_404(id)
    nome = (iscritto.NomeRagazzo or "").strip()
    cognome = (iscritto.CognomeRagazzo or "").strip()
    email = (iscritto.MailRicevuta or "").strip()

    if not email:
        return jsonify({"error": "Email ricevuta non disponibile per questo iscritto."}), 400
    if not _EMAIL_RE.match(email):
        return jsonify({"error": "L'email ricevuta salvata non e' valida."}), 400

    try:
        status_code = _invoke_resend_confirmation_flow(nome, cognome, email)
        return jsonify({"ok": True, "status": status_code})
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 502


@bp.route("/api/export/iscritti")
def export_iscritti_excel():
    evento = get_current_event_id()
    config = load_config()
    numero_settimane = int(config["settimane"]["numero_settimane"])

    iscritti = (Iscritto.query
                .options(joinedload(Iscritto.contabilita)
                         .selectinload(Contabilita.settimane))
                .filter_by(ID_Evento=evento)
                .order_by(Iscritto.CognomeRagazzo, Iscritto.NomeRagazzo)
                .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Presenze"
    ws.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="DCEAF7")
    header_fill = PatternFill("solid", fgColor="4F8FC7")
    presence_fill = PatternFill("solid", fgColor="D1FAE5")
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True, color="1F2937")
    presence_font = Font(bold=True, color="047857", size=13)
    thin_side = Side(style="thin", color="9FB3C8")
    medium_side = Side(style="medium", color="4F8FC7")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    base_headers = ["Cognome", "Nome", "Classe frequentata", "Squadra"]
    settimana_headers = [f"Sett. {i}" for i in range(1, numero_settimane + 1)]
    headers = base_headers + settimana_headers + ["Presente oggi"]
    num_cols = len(headers)

    ws["A1"] = "Data"
    ws["A1"].font = Font(bold=True, size=12, color="1F2937")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws["A1"].border = header_border

    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(num_cols)
    ws.merge_cells(f"B1:{last_col_letter}1")
    ws["B1"] = ""
    ws["B1"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["B1"].border = header_border
    ws["B1"].alignment = left

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = center

    base_widths = [24, 22, 24, 18]
    for i, w in enumerate(base_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(base_widths) + 1, len(base_widths) + len(settimana_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions[last_col_letter].width = 16

    presence_col_idx = num_cols
    week_col_start = len(base_headers) + 1
    week_col_end = len(base_headers) + len(settimana_headers)

    start_row = 4
    for row_idx, iscritto in enumerate(iscritti, start=start_row):
        presenze = _presence_by_week(iscritto, numero_settimane)
        base_values = [
            iscritto.CognomeRagazzo,
            iscritto.NomeRagazzo,
            iscritto.ClasseFrequentata or "",
            iscritto.Squadra or "",
        ]
        for col_idx, value in enumerate(base_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = left
            if col_idx in (1, 2):
                cell.font = label_font

        for w in range(1, numero_settimane + 1):
            col_idx = len(base_headers) + w
            value = "X" if presenze.get(w) else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center
            if value:
                cell.font = presence_font
                cell.fill = presence_fill

        cell = ws.cell(row=row_idx, column=presence_col_idx, value="")
        cell.border = thin_border
        cell.alignment = center

    last_row = max(start_row, start_row + len(iscritti) - 1)
    presence_col_letter = get_column_letter(presence_col_idx)
    validation = DataValidation(type="list", formula1='"Si,No"', allow_blank=True)
    validation.prompt = "Seleziona la presenza giornaliera"
    validation.promptTitle = "Presenza"
    ws.add_data_validation(validation)
    validation.add(f"{presence_col_letter}{start_row}:{presence_col_letter}{last_row}")

    for row_idx in range(1, last_row + 1):
        ws.row_dimensions[row_idx].height = 24

    ws.freeze_panes = f"{get_column_letter(week_col_start)}4"
    ws.auto_filter.ref = f"A3:{last_col_letter}{last_row}"

    # Aggiunge i 7 fogli del report completo (riusa la logica di genera_report.py)
    raw_conn = db.engine.raw_connection()
    try:
        cursor = raw_conn.cursor()
        _rpt.foglio_pranzo_sett1(wb, cursor)
        _rpt.foglio_allergie_animatori(wb, cursor)
        _rpt.foglio_allergie_cucina(wb, cursor)
        _rpt.foglio_magliette(wb, cursor)
        _rpt.foglio_distribuzione_magliette(wb, cursor)
        _rpt.foglio_no_foto(wb, cursor)
        _rpt.foglio_navetta(wb, cursor)
        _rpt.foglio_uscita_autonoma(wb, cursor)
        _rpt.foglio_pagamenti(wb, cursor)
        cursor.close()
    finally:
        raw_conn.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="report_oratorio.xlsx",
    )


def _calcola_assegnazioni_squadre(cursor, seed):
    """Esegue l'algoritmo di assegnazione in modo deterministico (dato il seed).
    Restituisce: (assegnazioni_dict, distribuzione_finale, tutti_iscritti, famiglia_di)."""
    _random.seed(seed)

    tutti = _aq.carica_iscritti(cursor, solo_senza_squadra=False)
    da_assegnare = [iso for iso in tutti
                    if not iso["squadra_attuale"]
                    or iso["squadra_attuale"].strip() == ""]

    famiglia_di, gruppi = _aq.identifica_famiglie(tutti)
    if any(iso["squadra_attuale"] for iso in tutti):
        contatori = _aq.carica_contatori_esistenti(cursor)
    else:
        contatori = {s: _defaultdict(lambda: _defaultdict(int))
                     for s in _aq.SQUADRE}

    assegnazioni = _aq.assegna(da_assegnare, contatori, famiglia_di, gruppi)
    return assegnazioni, contatori, tutti, famiglia_di


def _build_preview_payload(assegnazioni, contatori, tutti, famiglia_di, seed):
    iso_map = {iso["id"]: iso for iso in tutti}

    lista = sorted(
        [
            {
                "id": iso_id,
                "cognome": iso_map[iso_id]["cognome"],
                "nome": iso_map[iso_id]["nome"],
                "classe": iso_map[iso_id]["classe"],
                "squadra": sq,
            }
            for iso_id, sq in assegnazioni.items()
        ],
        key=lambda x: (x["squadra"], x["classe"] or "", x["cognome"] or ""),
    )

    distribuzione = {}
    for sq in _aq.SQUADRE:
        per_classe = {}
        tot_m = tot_f = 0
        for classe in _aq.CLASSE_ORDER:
            m = contatori[sq][classe]["M"]
            f = contatori[sq][classe]["F"]
            per_classe[classe] = {"M": m, "F": f}
            tot_m += m
            tot_f += f
        distribuzione[sq] = {
            "per_classe": per_classe,
            "totale_m": tot_m,
            "totale_f": tot_f,
            "totale": tot_m + tot_f,
        }

    collisioni = _defaultdict(list)
    for iso_id, sq in assegnazioni.items():
        fam = famiglia_di.get(iso_id)
        if fam:
            collisioni[(fam, sq)].append(iso_map[iso_id])
    collisioni_fratelli = [
        {
            "squadra": sq,
            "membri": [f"{m['nome']} {m['cognome']} ({m['classe']})" for m in membri],
        }
        for (fam, sq), membri in collisioni.items() if len(membri) > 1
    ]

    return {
        "seed": seed,
        "totale_da_assegnare": len(assegnazioni),
        "assegnazioni": lista,
        "distribuzione": distribuzione,
        "collisioni_fratelli": collisioni_fratelli,
        "classi_ordinate": _aq.CLASSE_ORDER,
        "squadre": _aq.SQUADRE,
    }


@bp.route("/api/squadre/preview", methods=["POST"])
def squadre_preview():
    """Calcola l'assegnazione squadre in dry-run e ritorna il piano + seed."""
    seed = _secrets.randbits(32)

    raw_conn = db.engine.raw_connection()
    try:
        cursor = raw_conn.cursor()
        assegnazioni, contatori, tutti, famiglia_di = _calcola_assegnazioni_squadre(cursor, seed)
        cursor.close()
    finally:
        raw_conn.close()

    if not assegnazioni:
        return jsonify({
            "totale_da_assegnare": 0,
            "messaggio": "Tutti gli iscritti hanno già una squadra assegnata.",
        })

    payload = _build_preview_payload(assegnazioni, contatori, tutti, famiglia_di, seed)
    return jsonify(payload)


@bp.route("/api/squadre/applica", methods=["POST"])
def squadre_applica():
    """Applica le assegnazioni squadre usando il seed dal preview."""
    data = request.get_json(silent=True) or {}
    seed = data.get("seed")
    if seed is None:
        return jsonify({"error": "Parametro 'seed' mancante."}), 400
    try:
        seed = int(seed)
    except (TypeError, ValueError):
        return jsonify({"error": "Parametro 'seed' non valido."}), 400

    raw_conn = db.engine.raw_connection()
    try:
        cursor = raw_conn.cursor()
        assegnazioni, _c, _t, _f = _calcola_assegnazioni_squadre(cursor, seed)
        cursor.close()
    finally:
        raw_conn.close()

    if not assegnazioni:
        return jsonify({"ok": True, "applicati": 0,
                        "messaggio": "Nessun iscritto da assegnare."})

    for iso_id, sq in assegnazioni.items():
        iscritto = Iscritto.query.get(iso_id)
        if iscritto:
            iscritto.Squadra = sq
    db.session.commit()

    return jsonify({"ok": True, "applicati": len(assegnazioni)})


def _update_iscritto_fields(iscritto, data):
    date_fields = {"DataNascitaRagazzo", "DataValidazione"}
    bool_fields = {"Navetta", "UscitaAutorizzata", "IscrizioneValidata", "MagliettaConsegnata"}
    skip = {"ID", "ID_Evento", "DataCreazione", "DataModifica", "ModificatoDa"} | _PRESENZA_FIELDS

    for col in Iscritto.__table__.columns:
        if col.name in skip or col.name not in data:
            continue
        val = data[col.name]
        if col.name in date_fields:
            val = _parse_iso_date(val)
        elif col.name in bool_fields:
            val = _coerce_bool(val)
        elif isinstance(val, str):
            val = val.strip() or None
        setattr(iscritto, col.name, val)
