#!/usr/bin/env python3
"""
Genera il report Excel per l'Oratorio Estivo.
Uso: python genera_report.py
Il file viene salvato nella stessa cartella come Report_OratorioEstivo.xlsx
"""

import os
import sys
from decimal import Decimal
from pathlib import Path

try:
    import pyodbc
except ImportError:
    sys.exit("Installa pyodbc: pip install pyodbc")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Installa openpyxl: pip install openpyxl")

from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

DB_SERVER = os.environ["DB_SERVER"]
DB_NAME = os.environ.get("DB_NAME", "oratorio-estivo")
DB_USER = os.environ["DB_USER"]
DB_PASSWORD = os.environ["DB_PASSWORD"]
DB_DRIVER = os.environ.get("DB_ODBC_DRIVER", "ODBC Driver 18 for SQL Server")
OUTPUT_PATH = BASE_DIR / "Report_OratorioEstivo.xlsx"

ALLERGIE_ESCLUSE = (
    "nessuna", "no", "nulla", "nessuno", "", "niente", "n/a", "/", "-", "//",
    ".", "no .", "nessun problema", "nessuna allergia", "nessuna nota",
    "nessun tipo di allergia o intolleranza", "nessuna allergia o intolleranza",
    "nessuna intolleranza", "no niente",
)

# ── Stili ──────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")

FILLS = {
    "red":    PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
    "blue":   PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
    "green":  PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
    "orange": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
    "purple": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
    "teal":   PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
    "brown":  PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid"),
}


def style_headers(ws, headers, fill_name="blue"):
    fill = FILLS[fill_name]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def cell(ws, row, col, value, center=False, wrap=False, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN_BORDER
    if center:
        c.alignment = CENTER
    elif wrap:
        c.alignment = WRAP
    else:
        c.alignment = Alignment(vertical="top")
    if bold:
        c.font = Font(bold=True)
    return c


def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = [
            max((len(line) for line in str(c.value).split("\n")), default=0)
            for c in col if c.value
        ]
        ws.column_dimensions[letter].width = min(max(max(lengths, default=min_w) + 2, min_w), max_w)


def si_no(val):
    return "Sì" if val else "No"


def has_real_allergia(val):
    if not val:
        return False
    return val.strip().lower() not in ALLERGIE_ESCLUSE


# ── Fogli ──────────────────────────────────────────────────────

def foglio_pranzo_sett1(wb, cursor):
    ws = wb.create_sheet("PRANZO Sett.1 (URGENTE)")
    ws.sheet_properties.tabColor = "C00000"
    style_headers(ws, ["#", "Cognome", "Nome", "Classe", "Squadra",
                        "Mattina", "Pranzo", "Pomeriggio"], "red")

    cursor.execute("""
        SELECT i.CognomeRagazzo, i.NomeRagazzo, i.ClasseFrequentata, i.Squadra,
               ps.Mattina, ps.Pranzo, ps.Pomeriggio
        FROM dbo.pagamenti_settimanali ps
        JOIN dbo.contabilita c ON ps.ID_Contabilita = c.ID
        JOIN dbo.iscritti i ON c.ID_Iscritto = i.ID
        WHERE ps.NumeroSettimana = 1 AND ps.Pranzo = 1
        ORDER BY i.ClasseFrequentata, i.CognomeRagazzo, i.NomeRagazzo
    """)
    rows = cursor.fetchall()
    for i, r in enumerate(rows, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2])
        cell(ws, rn, 5, r[3] or "—")
        cell(ws, rn, 6, si_no(r[4]), center=True)
        cell(ws, rn, 7, si_no(r[5]), center=True)
        cell(ws, rn, 8, si_no(r[6]), center=True)

    tr = len(rows) + 2
    cell(ws, tr, 2, "TOTALE PRANZO SETT. 1:", bold=True)
    cell(ws, tr, 3, len(rows), bold=True)
    auto_width(ws)
    print(f"  Pranzo Sett.1: {len(rows)} iscritti")


def foglio_allergie_animatori(wb, cursor):
    ws = wb.create_sheet("Allergie Animatori")
    ws.sheet_properties.tabColor = "ED7D31"
    style_headers(ws, ["#", "Cognome", "Nome", "Allergie / Intolleranze"], "orange")

    cursor.execute("""
        SELECT Cognome, Nome, AllergieIntolleranze
        FROM animatori.animatori
        ORDER BY Cognome
    """)
    idx = 0
    for r in cursor.fetchall():
        if not has_real_allergia(r[2]):
            continue
        idx += 1
        rn = idx + 1
        cell(ws, rn, 1, idx, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2], wrap=True)

    auto_width(ws)
    ws.column_dimensions["D"].width = 50
    print(f"  Allergie Animatori: {idx}")


def foglio_allergie_cucina(wb, cursor):
    ws = wb.create_sheet("Allergie Cucina")
    ws.sheet_properties.tabColor = "C00000"
    style_headers(ws, ["#", "Cognome", "Nome", "Classe", "Squadra",
                        "Allergie / Intolleranze"], "red")

    cursor.execute("""
        SELECT CognomeRagazzo, NomeRagazzo, ClasseFrequentata, Squadra,
               AllergieIntolleranze
        FROM dbo.iscritti
        ORDER BY ClasseFrequentata, CognomeRagazzo
    """)
    idx = 0
    for r in cursor.fetchall():
        if not has_real_allergia(r[4]):
            continue
        idx += 1
        rn = idx + 1
        cell(ws, rn, 1, idx, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2], center=True)
        cell(ws, rn, 5, r[3] or "—", center=True)
        cell(ws, rn, 6, r[4], wrap=True)

    auto_width(ws)
    ws.column_dimensions["F"].width = 55
    print(f"  Allergie Cucina: {idx}")


def foglio_magliette(wb, cursor):
    ws = wb.create_sheet("Magliette")
    ws.sheet_properties.tabColor = "548235"

    SQUADRA_COLORS = {
        "Blu":    PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
        "Rossa":  PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        "Verde":  PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
        "Gialla": PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid"),
    }

    taglia_order = [
        "7/8 anni", "9/10 anni", "11/12 anni", "13/14 anni",
        "S", "M", "L", "XL", "2XL",
    ]

    cursor.execute("""
        SELECT i.Squadra, i.TagliaMaglietta,
               COUNT(*) + COALESCE(SUM(c.NumeroMaglietteExtra), 0) AS totale
        FROM dbo.iscritti i
        LEFT JOIN dbo.contabilita c ON c.ID_Iscritto = i.ID
        WHERE i.TagliaMaglietta IS NOT NULL
          AND i.Squadra IS NOT NULL AND LTRIM(RTRIM(i.Squadra)) != ''
        GROUP BY i.Squadra, i.TagliaMaglietta
        ORDER BY i.Squadra, i.TagliaMaglietta
    """)
    rows = cursor.fetchall()
    taglie = sorted(
        {r[1] for r in rows},
        key=lambda x: taglia_order.index(x) if x in taglia_order else 99,
    )
    squadre = ["Blu", "Rossa", "Verde", "Gialla"]

    style_headers(ws, ["Squadra"] + taglie + ["TOTALE"], "green")

    data = {}
    for r in rows:
        data.setdefault(r[0], {})[r[1]] = r[2]

    grand_total = 0
    for si, squadra in enumerate(squadre, 1):
        rn = si + 1
        sq_cell = cell(ws, rn, 1, squadra, bold=True)
        sq_cell.fill = SQUADRA_COLORS.get(squadra, FILLS["blue"])
        sq_cell.font = Font(bold=True, color="FFFFFF")
        row_tot = 0
        for ti, taglia in enumerate(taglie, 2):
            v = data.get(squadra, {}).get(taglia, 0)
            cell(ws, rn, ti, v if v else "", center=True)
            row_tot += v
        cell(ws, rn, len(taglie) + 2, row_tot, center=True, bold=True)
        grand_total += row_tot

    tr = len(squadre) + 2
    cell(ws, tr, 1, "TOTALE", bold=True)
    for ti, taglia in enumerate(taglie, 2):
        col_tot = sum(data.get(s, {}).get(taglia, 0) for s in squadre)
        cell(ws, tr, ti, col_tot, center=True, bold=True)
    c = cell(ws, tr, len(taglie) + 2, grand_total, center=True, bold=True)
    c.font = Font(bold=True, size=12, color="C00000")

    auto_width(ws)
    print(f"  Magliette: {grand_total} totali, {len(squadre)} squadre x {len(taglie)} taglie")


def foglio_distribuzione_magliette(wb, cursor):
    """Elenco completo per consegna magliette: nome, cognome, squadra, classe,
    taglia, extra. Ordinato per squadra → classe → cognome."""
    ws = wb.create_sheet("Distrib. Magliette")
    ws.sheet_properties.tabColor = "70AD47"
    style_headers(ws, [
        "#", "Cognome", "Nome", "Squadra", "Classe", "Taglia", "Extra", "Consegnata",
    ], "green")

    cursor.execute("""
        SELECT i.CognomeRagazzo, i.NomeRagazzo, i.Squadra, i.ClasseFrequentata,
               i.TagliaMaglietta,
               COALESCE(c.NumeroMaglietteExtra, 0) AS Extra,
               i.MagliettaConsegnata
        FROM dbo.iscritti i
        LEFT JOIN dbo.contabilita c ON c.ID_Iscritto = i.ID
        ORDER BY
            CASE i.Squadra
                WHEN 'Blu' THEN 1 WHEN 'Rossa' THEN 2
                WHEN 'Verde' THEN 3 WHEN 'Gialla' THEN 4
                ELSE 99 END,
            i.ClasseFrequentata, i.CognomeRagazzo, i.NomeRagazzo
    """)
    rows = cursor.fetchall()
    tot_base = len(rows)
    tot_extra = 0
    tot_consegnate = 0
    for i, r in enumerate(rows, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2] or "—", center=True)
        cell(ws, rn, 5, r[3] or "—", center=True)
        cell(ws, rn, 6, r[4] or "—", center=True, bold=True)
        extra = int(r[5] or 0)
        c = cell(ws, rn, 7, extra if extra else "", center=True)
        if extra > 0:
            c.font = Font(bold=True, color="C00000")
        tot_extra += extra
        consegnata = bool(r[6])
        cc = cell(ws, rn, 8, "X" if consegnata else "", center=True)
        if consegnata:
            cc.font = Font(bold=True, color="047857")
            tot_consegnate += 1

    tr = len(rows) + 2
    cell(ws, tr, 2, "TOTALE", bold=True)
    cell(ws, tr, 6, f"{tot_base} base", center=True, bold=True)
    cell(ws, tr, 7, f"+ {tot_extra} extra", center=True, bold=True).font = \
        Font(bold=True, color="C00000")
    cell(ws, tr, 8, f"{tot_consegnate} ✓", center=True, bold=True).font = \
        Font(bold=True, color="047857")

    auto_width(ws)
    print(f"  Distrib. Magliette: {tot_base} base + {tot_extra} extra = "
          f"{tot_base + tot_extra} magliette totali ({tot_consegnate} consegnate)")


def foglio_no_foto(wb, cursor):
    ws = wb.create_sheet("No Foto")
    ws.sheet_properties.tabColor = "7030A0"
    style_headers(ws, ["#", "Cognome", "Nome", "Classe", "Squadra", "Note"], "purple")

    cursor.execute("""
        SELECT CognomeRagazzo, NomeRagazzo, ClasseFrequentata, Squadra, TerapieNote
        FROM dbo.iscritti
        WHERE LOWER(TerapieNote) LIKE '%foto%'
        ORDER BY ClasseFrequentata, CognomeRagazzo
    """)
    rows = cursor.fetchall()
    for i, r in enumerate(rows, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2], center=True)
        cell(ws, rn, 5, r[3] or "—", center=True)
        cell(ws, rn, 6, r[4])

    auto_width(ws)
    print(f"  No Foto: {len(rows)}")


def foglio_navetta(wb, cursor):
    ws = wb.create_sheet("Navetta")
    ws.sheet_properties.tabColor = "2E75B6"
    style_headers(ws, ["#", "Cognome", "Nome", "Classe", "Squadra"], "teal")

    cursor.execute("""
        SELECT CognomeRagazzo, NomeRagazzo, ClasseFrequentata, Squadra
        FROM dbo.iscritti
        WHERE Navetta = 1
        ORDER BY ClasseFrequentata, CognomeRagazzo
    """)
    rows = cursor.fetchall()
    for i, r in enumerate(rows, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2], center=True)
        cell(ws, rn, 5, r[3] or "—", center=True)

    auto_width(ws)
    print(f"  Navetta: {len(rows)}")


def foglio_uscita_autonoma(wb, cursor):
    ws = wb.create_sheet("Uscita Autonoma")
    ws.sheet_properties.tabColor = "C55A11"
    style_headers(ws, ["#", "Cognome", "Nome", "Classe", "Squadra"], "orange")

    cursor.execute("""
        SELECT CognomeRagazzo, NomeRagazzo, ClasseFrequentata, Squadra
        FROM dbo.iscritti
        WHERE UscitaAutorizzata = 1
        ORDER BY ClasseFrequentata, CognomeRagazzo
    """)
    rows = cursor.fetchall()
    for i, r in enumerate(rows, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2], center=True)
        cell(ws, rn, 5, r[3] or "—", center=True)

    auto_width(ws)
    print(f"  Uscita Autonoma: {len(rows)}")


def foglio_pagamenti(wb, cursor):
    ws = wb.create_sheet("Pagamenti")
    ws.sheet_properties.tabColor = "843C0C"
    style_headers(ws, [
        "#", "Cognome", "Nome", "Classe", "Settimana",
        "Mattina", "Pranzo", "Pomeriggio", "Gita", "Imp. Gita",
        "Totale", "Pagato", "Data Pag.",
    ], "brown")

    cursor.execute("""
        SELECT i.CognomeRagazzo, i.NomeRagazzo, i.ClasseFrequentata,
               ps.NumeroSettimana, ps.Mattina, ps.Pranzo, ps.Pomeriggio,
               ps.GitaSettimana, ps.ImportoGita,
               COALESCE(ps.TotaleManuale, ps.Totale) AS TotaleEffettivo,
               ps.Pagato, ps.DataPagamento
        FROM dbo.pagamenti_settimanali ps
        JOIN dbo.contabilita c ON ps.ID_Contabilita = c.ID
        JOIN dbo.iscritti i ON c.ID_Iscritto = i.ID
        WHERE ps.Pagato = 1
        ORDER BY i.ClasseFrequentata, i.CognomeRagazzo, i.NomeRagazzo,
                 ps.NumeroSettimana
    """)
    rows = cursor.fetchall()

    tot_m, tot_p, tot_pm, tot_eur = 0, 0, 0, Decimal("0")
    for i, r in enumerate(rows, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2], center=True)
        cell(ws, rn, 5, f"Sett. {r[3]}", center=True)
        cell(ws, rn, 6, si_no(r[4]), center=True)
        cell(ws, rn, 7, si_no(r[5]), center=True)
        cell(ws, rn, 8, si_no(r[6]), center=True)
        cell(ws, rn, 9, si_no(r[7]), center=True)
        cell(ws, rn, 10, float(r[8] or 0), center=True)
        cell(ws, rn, 11, float(r[9] or 0), center=True)
        cell(ws, rn, 12, si_no(r[10]), center=True)
        cell(ws, rn, 13, str(r[11])[:10] if r[11] else "—", center=True)

        if r[4]: tot_m += 1
        if r[5]: tot_p += 1
        if r[6]: tot_pm += 1
        if r[9]: tot_eur += r[9]

    tr = len(rows) + 2
    for c in range(1, 14):
        ws.cell(row=tr, column=c).border = THIN_BORDER
    cell(ws, tr, 2, "TOTALI", bold=True)
    cell(ws, tr, 6, tot_m, center=True, bold=True)
    cell(ws, tr, 7, tot_p, center=True, bold=True)
    cell(ws, tr, 8, tot_pm, center=True, bold=True)
    tc = cell(ws, tr, 11, float(tot_eur), center=True, bold=True)
    tc.font = Font(bold=True, color="C00000", size=12)
    tc.number_format = '#,##0.00 €'

    auto_width(ws)
    print(f"  Pagamenti: {len(rows)} righe — "
          f"matt {tot_m} | pranzo {tot_p} | pom {tot_pm} | €{tot_eur:,.2f}")


# ── Main ───────────────────────────────────────────────────────

def main():
    print(f"Connessione a {DB_SERVER} / {DB_NAME} ...")
    conn = pyodbc.connect(
        f"DRIVER={{{DB_DRIVER}}};"
        f"SERVER={DB_SERVER};DATABASE={DB_NAME};"
        f"UID={DB_USER};PWD={DB_PASSWORD};"
        "Encrypt=yes;TrustServerCertificate=no"
    )
    cursor = conn.cursor()
    print("Connesso.\n")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("Generazione fogli:")
    foglio_pranzo_sett1(wb, cursor)
    foglio_allergie_animatori(wb, cursor)
    foglio_allergie_cucina(wb, cursor)
    foglio_magliette(wb, cursor)
    foglio_distribuzione_magliette(wb, cursor)
    foglio_no_foto(wb, cursor)
    foglio_navetta(wb, cursor)
    foglio_uscita_autonoma(wb, cursor)
    foglio_pagamenti(wb, cursor)

    wb.save(str(OUTPUT_PATH))
    conn.close()
    print(f"\nReport salvato in: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
