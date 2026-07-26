#!/usr/bin/env python3
"""
Genera il report Excel per gli Animatori dell'Oratorio Estivo.
Uso: python genera_report_animatori.py
Il file viene salvato come Report_Animatori.xlsx

Le stesse funzioni-foglio sono importate dall'endpoint Flask della webapp,
così la logica resta unica.
"""

import os
import sys
from decimal import Decimal
from pathlib import Path

try:
    import pyodbc
except ImportError:
    sys.exit("Installa pyodbc: pip install pyodbc")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Installa openpyxl: pip install openpyxl")

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = BASE_DIR / "Report_Animatori.xlsx"

ALLERGIE_ESCLUSE = (
    "nessuna", "no", "nulla", "nessuno", "", "niente", "n/a", "/", "-", "//",
    ".", "no .", "nessun problema", "nessuna allergia", "nessuna nota",
    "nessun tipo di allergia o intolleranza", "nessuna allergia o intolleranza",
    "nessuna intolleranza", "no niente",
)

# ── Stili ──────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")

FILLS = {
    "red":    PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
    "blue":   PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
    "green":  PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
    "orange": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
    "teal":   PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
}
PRESENCE_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")


def style_headers(ws, headers, fill_name="blue"):
    fill = FILLS[fill_name]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER


def cell(ws, row, col, value, center=False, wrap=False, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN_BORDER
    if center:
        c.alignment = CENTER
    elif wrap:
        c.alignment = WRAP
    else:
        c.alignment = Alignment(vertical="top")
    if bold:
        c.font = Font(bold=True)
    return c


def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = [
            max((len(line) for line in str(c.value).split("\n")), default=0)
            for c in col if c.value
        ]
        ws.column_dimensions[letter].width = min(max(max(lengths, default=min_w) + 2, min_w), max_w)


def si_no(val):
    return "Sì" if val else "No"


def has_real_allergia(val):
    if not val:
        return False
    return val.strip().lower() not in ALLERGIE_ESCLUSE


def _numero_settimane(cursor):
    cursor.execute("""
        SELECT TOP 1 NumeroSettimane FROM animatori.eventi_animatori
        WHERE Attivo = 1 ORDER BY Anno DESC
    """)
    r = cursor.fetchone()
    if r and r[0]:
        return int(r[0])
    cursor.execute("SELECT ISNULL(MAX(NumeroSettimana), 5) FROM animatori.disponibilita_animatori")
    return int(cursor.fetchone()[0] or 5)


# ── Fogli ──────────────────────────────────────────────────────

def foglio_anagrafica(wb, cursor):
    """Elenco animatori con disponibilità settimanale (X) + colonna presenza manuale."""
    ws = wb.create_sheet("Anagrafica")
    ws.sheet_properties.tabColor = "2F5496"

    n_sett = _numero_settimane(cursor)
    sett_headers = [f"Sett. {i}" for i in range(1, n_sett + 1)]
    headers = ["#", "Cognome", "Nome", "Cellulare", "Magg."] + sett_headers + ["Presente oggi"]
    style_headers(ws, headers, "blue")

    cursor.execute("""
        SELECT a.ID, a.Cognome, a.Nome, a.Cellulare, a.Maggiorenne
        FROM animatori.animatori a
        ORDER BY a.Cognome, a.Nome
    """)
    animatori = cursor.fetchall()

    # disponibilità per animatore/settimana
    cursor.execute("""
        SELECT ID_Animatore, NumeroSettimana, Disponibile
        FROM animatori.disponibilita_animatori
    """)
    disp = {}
    for r in cursor.fetchall():
        disp.setdefault(r[0], {})[int(r[1])] = bool(r[2])

    base_cols = 5  # #, Cognome, Nome, Cellulare, Magg.
    for i, a in enumerate(animatori, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, a[1])
        cell(ws, rn, 3, a[2])
        cell(ws, rn, 4, a[3] or "—", center=True)
        cell(ws, rn, 5, si_no(a[4]), center=True)
        for w in range(1, n_sett + 1):
            col = base_cols + w
            disponibile = disp.get(a[0], {}).get(w, False)
            c = cell(ws, rn, col, "X" if disponibile else "", center=True)
            if disponibile:
                c.font = Font(bold=True, color="047857", size=12)
                c.fill = PRESENCE_FILL
        cell(ws, rn, base_cols + n_sett + 1, "", center=True)

    last_row = max(2, len(animatori) + 1)
    pres_col = base_cols + n_sett + 1
    pres_letter = get_column_letter(pres_col)
    dv = DataValidation(type="list", formula1='"Si,No"', allow_blank=True)
    dv.promptTitle = "Presenza"
    dv.prompt = "Seleziona la presenza giornaliera"
    ws.add_data_validation(dv)
    dv.add(f"{pres_letter}2:{pres_letter}{last_row}")

    ws.freeze_panes = f"{get_column_letter(base_cols + 1)}2"
    auto_width(ws)
    print(f"  Anagrafica: {len(animatori)} animatori, {n_sett} settimane")


def foglio_magliette(wb, cursor):
    """Distribuzione magliette/pantaloncini con flag consegna."""
    ws = wb.create_sheet("Magliette")
    ws.sheet_properties.tabColor = "548235"
    style_headers(ws, [
        "#", "Cognome", "Nome", "Maglietta", "Pantaloncini", "Extra", "Consegnata",
    ], "green")

    cursor.execute("""
        SELECT a.Cognome, a.Nome, a.TagliaMaglietta, a.TagliaPantaloncini,
               COALESCE(c.NumeroMaglietteExtra, 0) AS Extra,
               a.MagliettaConsegnata
        FROM animatori.animatori a
        LEFT JOIN animatori.contributi_animatori c ON c.ID_Animatore = a.ID
        ORDER BY a.Cognome, a.Nome
    """)
    rows = cursor.fetchall()
    tot_extra = 0
    tot_consegnate = 0
    for i, r in enumerate(rows, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2] or "—", center=True, bold=True)
        cell(ws, rn, 5, r[3] or "—", center=True, bold=True)
        extra = int(r[4] or 0)
        c = cell(ws, rn, 6, extra if extra else "", center=True)
        if extra > 0:
            c.font = Font(bold=True, color="C00000")
        tot_extra += extra
        consegnata = bool(r[5])
        cc = cell(ws, rn, 7, "X" if consegnata else "", center=True)
        if consegnata:
            cc.font = Font(bold=True, color="047857")
            tot_consegnate += 1

    tr = len(rows) + 2
    cell(ws, tr, 2, "TOTALE", bold=True)
    cell(ws, tr, 4, f"{len(rows)} maglie", center=True, bold=True)
    cell(ws, tr, 6, f"+ {tot_extra} extra", center=True, bold=True).font = \
        Font(bold=True, color="C00000")
    cell(ws, tr, 7, f"{tot_consegnate} ✓", center=True, bold=True).font = \
        Font(bold=True, color="047857")

    auto_width(ws)
    print(f"  Magliette: {len(rows)} animatori + {tot_extra} extra "
          f"({tot_consegnate} consegnate)")


def foglio_allergie(wb, cursor):
    ws = wb.create_sheet("Allergie")
    ws.sheet_properties.tabColor = "ED7D31"
    style_headers(ws, ["#", "Cognome", "Nome", "Allergie / Intolleranze"], "orange")

    cursor.execute("""
        SELECT Cognome, Nome, AllergieIntolleranze
        FROM animatori.animatori
        ORDER BY Cognome, Nome
    """)
    idx = 0
    for r in cursor.fetchall():
        if not has_real_allergia(r[2]):
            continue
        idx += 1
        rn = idx + 1
        cell(ws, rn, 1, idx, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2], wrap=True)

    auto_width(ws)
    ws.column_dimensions["D"].width = 50
    print(f"  Allergie: {idx}")


def foglio_navetta(wb, cursor):
    ws = wb.create_sheet("Navetta")
    ws.sheet_properties.tabColor = "2E75B6"
    style_headers(ws, ["#", "Cognome", "Nome", "Cellulare"], "teal")

    cursor.execute("""
        SELECT Cognome, Nome, Cellulare
        FROM animatori.animatori
        WHERE Navetta = 1
        ORDER BY Cognome, Nome
    """)
    rows = cursor.fetchall()
    for i, r in enumerate(rows, 1):
        rn = i + 1
        cell(ws, rn, 1, i, center=True)
        cell(ws, rn, 2, r[0])
        cell(ws, rn, 3, r[1])
        cell(ws, rn, 4, r[2] or "—", center=True)

    auto_width(ws)
    print(f"  Navetta: {len(rows)}")


# ── Main ───────────────────────────────────────────────────────

def main():
    from dotenv import load_dotenv

    # Lo script standalone vive in animatori_azure/webapp_azure/ ; il .env sta nella root del repo
    for candidate in (BASE_DIR / ".env", BASE_DIR.parent.parent / ".env"):
        if candidate.exists():
            load_dotenv(candidate)
            break

    db_server = os.environ["DB_SERVER"]
    db_name = os.environ.get("DB_NAME", "oratorio-estivo")
    db_user = os.environ["DB_USER"]
    db_password = os.environ["DB_PASSWORD"]
    db_driver = os.environ.get("DB_ODBC_DRIVER", "ODBC Driver 18 for SQL Server")

    print(f"Connessione a {db_server} / {db_name} ...")
    conn = pyodbc.connect(
        f"DRIVER={{{db_driver}}};"
        f"SERVER={db_server};DATABASE={db_name};"
        f"UID={db_user};PWD={db_password};"
        "Encrypt=yes;TrustServerCertificate=no"
    )
    cursor = conn.cursor()
    print("Connesso.\n")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("Generazione fogli:")
    foglio_anagrafica(wb, cursor)
    foglio_magliette(wb, cursor)
    foglio_allergie(wb, cursor)
    foglio_navetta(wb, cursor)

    wb.save(str(OUTPUT_PATH))
    conn.close()
    print(f"\nReport salvato in: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
